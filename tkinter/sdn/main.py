## Criado por José de Senna Pereira Neto.
## Última modificação em 16/08/2024.

import tkinter as tk
import numpy as np
import xlsxwriter as xw
import xlrd as xd
import openpyxl as ox

from tkinter import filedialog, messagebox

## VARIÁVEIS GLOBAIS ################################################################################################

janela = tk.Tk()

arquivoDeEntrada = tk.StringVar()
arquivoDeSaida = tk.StringVar()
porcentagemRequerida = tk.Entry(janela)

## MANIPULAÇÃO DA PLANILHA ##########################################################################################

class FuncoesDeManipulacaoDaPlanilha():

    # ---------------------------------------------------------------------------------------------------------------
    # Recebe uma base de dados em excel e extrai o conteúdo linha a linha, armazenando em um dicionário.
    # Suporta apenas o formato *.xls e *.xlsx.
    # ---------------------------------------------------------------------------------------------------------------

    def criarListaDeEntrada(self, baseDeDados):

        if baseDeDados.endswith(".xls"):

            workbook = xd.open_workbook(baseDeDados)
            worksheet = workbook.sheet_by_index(0)

            listaDeNomesDeEntrada = {
                linha: (
                    worksheet.cell_value(linha, 0),
                    worksheet.cell_value(linha, 1),
                    worksheet.cell_value(linha, 2),
                    worksheet.cell_value(linha, 3),
                    worksheet.cell_value(linha, 4),
                    worksheet.cell_value(linha, 5),
                    worksheet.cell_value(linha, 6),
                    worksheet.cell_value(linha, 7),
                    worksheet.cell_value(linha, 8),
                    worksheet.cell_value(linha, 9),
                    worksheet.cell_value(linha, 10)
                )
                for linha in range(2, worksheet.nrows)
            }

        elif baseDeDados.endswith(".xlsx"):

            workbook = ox.load_workbook(baseDeDados)
            worksheet = workbook.active

            listaDeNomesDeEntrada = {
                linha: (
                    worksheet.cell(row=linha+1, column=1).value,
                    worksheet.cell(row=linha+1, column=2).value,
                    worksheet.cell(row=linha+1, column=3).value,
                    worksheet.cell(row=linha+1, column=4).value,
                    worksheet.cell(row=linha+1, column=5).value,
                    worksheet.cell(row=linha+1, column=6).value,
                    worksheet.cell(row=linha+1, column=7).value,
                    worksheet.cell(row=linha+1, column=8).value,
                    worksheet.cell(row=linha+1, column=9).value,
                    worksheet.cell(row=linha+1, column=10).value,
                    worksheet.cell(row=linha+1, column=11).value
                )
                for linha in range(1, worksheet.max_row)
            }

        else:

            raise ValueError(
                "Formato de arquivo não suportado. Por favor, selecione um arquivo .xls ou .xlsx."
            )

        return listaDeNomesDeEntrada

    # ---------------------------------------------------------------------------------------------------------------
    # Recebe um dicionário de tuplas e sorteia o conteúdo aleatoriamente com base num percentual requerido.
    # Retorna um dicionáio com as tuplas sorteadas.
    # ---------------------------------------------------------------------------------------------------------------

    def criarListaDeSaida(self, listaDeNomesDeEntrada, porcentagemDeNomesRequerida):

        listaDeNomes = list(listaDeNomesDeEntrada.keys())
        numeroDeNomesRequerido = int((len(listaDeNomes)*porcentagemDeNomesRequerida)/100)
        listaDeChaves = sorted(np.random.choice(listaDeNomes, numeroDeNomesRequerido, replace=False))

        listaDeNomesDeSaida = {
            chave: listaDeNomesDeEntrada[chave]
            for chave in listaDeChaves
        }

        return listaDeNomesDeSaida
    
    # ---------------------------------------------------------------------------------------------------------------
    # Recebe um dicionário de tuplas e um nome de arquivo e escreve os dados das tuplas no arquivo.
    # Armazena o processamento em uma planilha excel criada linha a linha proceduralmente.
    # ---------------------------------------------------------------------------------------------------------------
    
    def criarPlanilhaDeSaida(self, listaDeNomesDeSaida, nomeDaplanilhaDeSaida):

        workbook = xw.Workbook(nomeDaplanilhaDeSaida)           # Cria a pasta de trabalho
        worksheet = workbook.add_worksheet("Lista de Nomes")    # Cria a planilha

        # Seção para definir os cabeçalhos das colunas da planilha

        listaDeCabecalhosDaPlanilha = (
            "Nome",
            "Função",
            "CPF",
            "Data de Nascimento",
            "RG",
            "Data de Admissão",
            "Telefone",
            "Filial",
            "Matrícula",
            "E-mail",
            "Contrato",
        )

        # Seção para criar os formatos de célula utilizados na planilha

        formatoDeData = workbook.add_format({'num_format': 'dd/mm/yyyy'})

        # Seção para definir a largura e a formatação das colunas da planilha

        worksheet.set_column('A:A', 50)                         # Nome
        worksheet.set_column('B:B', 50)                         # Função
        worksheet.set_column('C:C', 20)                         # CPF
        worksheet.set_column('D:D', 20, formatoDeData)          # Data de Nascimento
        worksheet.set_column('E:E', 20)                         # RG
        worksheet.set_column('F:F', 20, formatoDeData)          # Data de Admissão
        worksheet.set_column('G:G', 20)                         # Telefone
        worksheet.set_column('H:H', 20)                         # Filial
        worksheet.set_column('I:I', 20)                         # Matrícula
        worksheet.set_column('J:J', 50)                         # E-mail
        worksheet.set_column('K:K', 50)                         # Contrato

        # Seção para escrever os dados do cabeçalho e do dicionário de tuplas na planilha

        worksheet.write_row(0, 0, listaDeCabecalhosDaPlanilha)        
        for linha, (chave, valores) in enumerate(listaDeNomesDeSaida.items(), start=1):
            worksheet.write_row(linha, 0, valores)

        workbook.close()                                        # Fecha a planilha

## FUNÇÕES DA INTERFACE GRÁFICA #####################################################################################

class FuncoesDaInterfaceGrafica(FuncoesDeManipulacaoDaPlanilha):

    # ---------------------------------------------------------------------------------------------------------------
    # Abre uma caixa de diálogo para navegar até uma planilha excel.
    # Armazena o endereço na variável global arquivoDeEntrada.
    # ---------------------------------------------------------------------------------------------------------------

    def selecionarArquivo(self):

        localizacaoDoArquivo = filedialog.askopenfilename(
            title = "Selecione o arquivo Excel",
            filetypes = [
                ("Pasta de Trabalho do Excel", "*.xlsx"),
                ("Pasta de Trabalho do Excel 97-2003", "*.xls"),
                ("Todos os arquivos", "*.*")
            ]
        )

        arquivoDeEntrada.set(localizacaoDoArquivo)

    # ---------------------------------------------------------------------------------------------------------------
    # Abre uma caixa de diálogo para navegar até a localização e definir um nome para uma planilha excel.
    # Armazena o nome e o endereço na variável global arquivoDeSaida.
    # ---------------------------------------------------------------------------------------------------------------

    def salvarArquivo(self):

        localizacaoDoArquivo = filedialog.asksaveasfilename(
            title = "Salvar Como",
            defaultextension=".xlsx",
            filetypes = [
                ("Pasta de Trabalho do Excel", "*.xlsx"),
                ("Todos os arquivos", "*.*")
            ]
        )

        arquivoDeSaida.set(localizacaoDoArquivo)
    
    # ---------------------------------------------------------------------------------------------------------------
    # Chama as funções de manipulação da planilha, utilizando as variáveis globais como parâmetros de execução.
    # Prevê os erros de não informar os arquivos de entrada e saída.
    # Informa o erro Python caso a criação da planilha não ocorra com sucesso.
    # ---------------------------------------------------------------------------------------------------------------

    def executarAplicacao(self):

        entrada = arquivoDeEntrada.get()
        saida = arquivoDeSaida.get()
        porcentagem = int(porcentagemRequerida.get())

        if not entrada or not saida:
            messagebox.showerror("Erro", "Por favor, selecione os arquivos de entrada e saída.")
            return
        
        try:
            listaDeEntrada = self.criarListaDeEntrada(entrada)
            listaDeSaida = self.criarListaDeSaida(listaDeEntrada, porcentagem)
            self.criarPlanilhaDeSaida(listaDeSaida, saida)
            messagebox.showinfo("Sucesso", f"Arquivo salvo com sucesso em: {saida}")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

## APLICAÇÃO ########################################################################################################

class Aplicacao(FuncoesDaInterfaceGrafica):

    # ---------------------------------------------------------------------------------------------------------------
    # Cria a janela principal e chama os elementos necessários.
    # ---------------------------------------------------------------------------------------------------------------

    def __init__(self):

        self.janela = janela
        self.formatacaoDaJanela()
        self.configuracaoDaTelaPrincipal()
        self.elementosDaTelaPrincipal()
        self.elementosDoFrameDeEntrada()
        self.elementosDoFrameDeSaida()
        self.janela.mainloop()

    # ---------------------------------------------------------------------------------------------------------------
    # Define as fontes e as cores dos elementos da aplicação.
    # ---------------------------------------------------------------------------------------------------------------

    def formatacaoDaJanela(self):

        self.corDaTela="#52536e"
        self.corDosFrames="#747591"
        self.corDasBordas="#8586a3"
        self.corDosBotoes="#9697b4"
        self.corDasLabels="#747591"
        self.corDasEntrys="#9697b4"
        self.corDaFonte="white"
        
        tipoDeFonte="calibri"
        tamanhoDaFonte=12
        formatacaoDaFonte="bold"
        self.fonte=(tipoDeFonte, tamanhoDaFonte, formatacaoDaFonte)

    # ---------------------------------------------------------------------------------------------------------------
    # Define o nome e o tamanho da janela principal.
    # ---------------------------------------------------------------------------------------------------------------

    def configuracaoDaTelaPrincipal(self):

        self.janela.title("Sorteador de Nomes")
        self.janela.configure(background=self.corDaTela)
        self.janela.geometry("800x600")
        self.janela.resizable(False, False)

    # ---------------------------------------------------------------------------------------------------------------
    # Define o nome e o posicionamento dos elementos que pertencem diretamente a janela principal.
    # Engloba o botão para executar o sorteio de nomes.
    # ---------------------------------------------------------------------------------------------------------------

    def elementosDaTelaPrincipal(self):

        self.nomeDaAplicacao = tk.Label(
            self.janela,
            text="Sorteador de Nomes",
            font=self.fonte,
            fg=self.corDaFonte,
            bg=self.corDaTela
        )

        self.nomeDaAplicacao.place(
            relx=0.05,
            rely=0.05,
        )

        self.descricaoDaAplicacao = tk.Label(
            self.janela,
            text="Sorteia uma lista de nomes a partir de uma base de dados Excel",
            font=self.fonte,
            fg=self.corDaFonte,
            bg=self.corDaTela
        )

        self.descricaoDaAplicacao.place(
            relx=0.05,
            rely=0.1,
        )

        self.textoDaPorcentagem = tk.Label(
            self.janela,
            text="Percentual Requerido:",
            font=self.fonte,
            fg=self.corDaFonte,
            bg=self.corDaTela
        )

        self.textoDaPorcentagem.place(
            relx=0.05,
            rely=0.81
        )

        porcentagemRequerida.place(
            relx=0.28,
            rely=0.8,
            relwidth=0.05,
            relheight=0.05
        )

        self.botaoDeExecucao = tk.Button(
            self.janela,
            text="Executar",
            font=self.fonte,
            fg=self.corDaFonte,
            bd=2,
            bg=self.corDosBotoes,
            command=self.executarAplicacao
        )

        self.botaoDeExecucao.place(
            relx=0.75,
            rely=0.8,
            relwidth=0.2,
            relheight=0.05
        )

    # ---------------------------------------------------------------------------------------------------------------
    # Define os nomes e os posicionamentos dos elementos que pertencem ao frame superior.
    # Engloba o botão para localizar o arquivo excel.
    # ---------------------------------------------------------------------------------------------------------------

    def elementosDoFrameDeEntrada(self):

        self.frameDeEntrada = tk.Frame(
            self.janela,
            bd=4,
            bg=self.corDosFrames,
            highlightbackground=self.corDasBordas,
            highlightthickness=4
        )

        self.frameDeEntrada.place(
            relx=0.05,
            rely=0.2,
            relwidth=0.9,
            relheight=0.25
        )

        self.textoDeEntrada = tk.Label(
            self.frameDeEntrada,
            text="Arquivo de Entrada:",
            font=self.fonte,
            fg=self.corDaFonte,
            bg=self.corDasLabels
        )

        self.textoDeEntrada.place(
            relx=0.01,
            rely=0.4,
        )

        self.caixaDeEntrada = tk.Entry(
            self.frameDeEntrada,
            textvariable=arquivoDeEntrada,
            font=self.fonte,
            fg=self.corDaFonte,
            bg=self.corDasEntrys
        )

        self.caixaDeEntrada.place(
            relx=0.25,
            rely=0.35,
            relwidth=0.5,
            relheight=0.25
        )

        self.botaoDeEntrada = tk.Button(
            self.frameDeEntrada,
            text="Localizar Arquivo",
            font=self.fonte,
            fg=self.corDaFonte,
            bd=2,
            bg=self.corDosBotoes,
            command=self.selecionarArquivo
        )

        self.botaoDeEntrada.place(
            relx=0.78,
            rely=0.35,
            relwidth=0.2,
            relheight=0.25
        )

    # ---------------------------------------------------------------------------------------------------------------
    # Define os nomes e os posicionamentos dos elementos que pertencem ao frame inferior.
    # Engloba o botão para salvar o arquivo excel.
    # ---------------------------------------------------------------------------------------------------------------

    def elementosDoFrameDeSaida(self):

        self.frameDeSaida = tk.Frame(
            self.janela,
            bd=4,
            bg=self.corDosFrames,
            highlightbackground=self.corDasBordas,
            highlightthickness=4
        )

        self.frameDeSaida.place(
            relx=0.05,
            rely=0.5,
            relwidth=0.9,
            relheight=0.25
        )

        self.textoDeSaida = tk.Label(
            self.frameDeSaida,
            text="Arquivo de Saída:",
            font=self.fonte,
            fg=self.corDaFonte,
            bg=self.corDasLabels
        )

        self.textoDeSaida.place(
            relx=0.01,
            rely=0.4,
        )

        self.caixaDeSaida = tk.Entry(
            self.frameDeSaida,
            textvariable=arquivoDeSaida,
            font=self.fonte,
            fg=self.corDaFonte,
            bg=self.corDasEntrys
        )

        self.caixaDeSaida.place(
            relx=0.25,
            rely=0.35,
            relwidth=0.5,
            relheight=0.25
        )

        self.botaoDeSaida = tk.Button(
            self.frameDeSaida,
            text="Salvar Como",
            font=self.fonte,
            fg=self.corDaFonte,
            bd=2,
            bg=self.corDosBotoes,
            command=self.salvarArquivo
        )

        self.botaoDeSaida.place(
            relx=0.78,
            rely=0.35,
            relwidth=0.2,
            relheight=0.25
        )

#####################################################################################################################

Aplicacao()